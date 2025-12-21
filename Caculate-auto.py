#!/usr/bin/env python3
"""
Caculate-auto.py

Tính số tiền một người (mặc định: Quân) cần trả dựa trên file CSV trong thư mục Transactions.

Chú ý: Google Sheets support đã bị loại bỏ theo yêu cầu.
"""

from decimal import Decimal
import csv
import os
import re
import argparse
from typing import Optional, List
from datetime import datetime, timedelta
import json
import sys

# optional libraries for exports
try:
	from openpyxl import Workbook
	HAS_OPENPYXL = True
except Exception:
	HAS_OPENPYXL = False

try:
	from reportlab.lib.pagesizes import A4
	from reportlab.pdfgen import canvas
	HAS_REPORTLAB = True
except Exception:
	HAS_REPORTLAB = False

TRANSACTIONS_DIR = 'Transactions'
SPLIT_RATIO = Decimal('0.5')

def parse_decimal(s: str) -> Decimal:
	try:
		if s is None or str(s).strip() == '':
			return Decimal('0')
		return Decimal(str(s))
	except Exception:
		m = re.search(r'-?[0-9]+', str(s) or '')
		if m:
			return Decimal(m.group(0))
		return Decimal('0')

def contains_person(text: str, person: str) -> bool:
	if not text:
		return False
	return person.lower() in text.lower()

def extract_amount_from_note(note: str) -> Optional[Decimal]:
	if not note:
		return None
	lines = [l.strip() for l in re.split(r'[\r\n]+', note) if l.strip()]
	for line in lines:
		txt = line.replace('\xa0', ' ')
		m = re.search(r'([0-9]+(?:[.,][0-9]{0,3})?)\s*[kK]\b', txt)
		if m:
			num = m.group(1).replace('.', '').replace(',', '')
			try:
				return Decimal(num) * Decimal('1000')
			except Exception:
				continue
		m = re.search(r'([0-9]{1,3}(?:[.,][0-9]{3})+|[0-9]+)\b', txt)
		if m:
			num = m.group(1).replace('.', '').replace(',', '')
			try:
				return Decimal(num)
			except Exception:
				continue
	return None

def list_transaction_files(base: str) -> List[str]:
	d = os.path.join(base, TRANSACTIONS_DIR)
	if not os.path.isdir(d):
		return []
	files = [f for f in os.listdir(d) if os.path.isfile(os.path.join(d, f))]
	files.sort()
	return files

def compute_summary(csv_path: str, person: str = 'Quân', paid_on: Optional[str] = None, start: Optional[str] = None):
	"""
	Compute summary from a CSV file path.

	Returns a dict with keys: totals, shared_rows, applied_payments, unapplied_payments.
	"""
	total_shared = Decimal('0')
	total_paid_by_person = Decimal('0')

	shared_rows = []
	payment_rows = []
	candidate_shared = []
	payments_by_person = []

	with open(csv_path, encoding='utf-8-sig', newline='') as f:
		reader = csv.DictReader(f)
		for row in reader:
			amt = parse_decimal(row.get('Amount', '0'))
			category = (row.get('Category name') or '')
			note = (row.get('Note') or '')
			date_str = row.get('Date') or ''
			try:
				dt = datetime.fromisoformat(date_str)
			except Exception:
				try:
					dt = datetime.fromisoformat(date_str.split('+')[0])
				except Exception:
					dt = None

			if amt > 0 and contains_person(note, person):
				payments_by_person.append(dt)
				payment_rows.append((date_str, dt, category, note, amt))

			if contains_person(category, person) or contains_person(note, person):
				if amt < 0:
					explicit = extract_amount_from_note(note)
					if explicit is not None and explicit > 0:
						share = explicit
						reason = 'explicit_in_note'
					else:
						share = (abs(amt) * SPLIT_RATIO).quantize(Decimal('1'))
						reason = 'split_ratio'
					candidate_shared.append((date_str, dt, category, note, amt, share, reason))

	cutoff_date = None
	if start:
		try:
			start_dt = datetime.strptime(start, '%Y-%m-%d').date()
			cutoff_date = start_dt
		except Exception:
			cutoff_date = None
	elif paid_on:
		try:
			paid_on_dt = datetime.strptime(paid_on, '%Y-%m-%d').date()
			cutoff_date = paid_on_dt + timedelta(days=1)
		except Exception:
			cutoff_date = None
	else:
		# Auto-detect if possible
		valid_payments = [p for p in payments_by_person if p is not None]
		if valid_payments:
			last_payment_dt = max(valid_payments)
			cutoff_date = (last_payment_dt.date() + timedelta(days=1))

	applied_payment_rows = []
	unapplied_payment_rows = []
	total_paid_by_person = Decimal('0')
	for _date_str, _dt, _cat, _note, amt in payment_rows:
		include_payment = True
		if cutoff_date:
			if _dt is None:
				include_payment = False
			else:
				include_payment = _dt.date() >= cutoff_date
		if include_payment:
			try:
				total_paid_by_person += Decimal(amt)
			except Exception:
				total_paid_by_person += parse_decimal(amt)
			applied_payment_rows.append((_date_str, _dt, _cat, _note, amt))
		else:
			unapplied_payment_rows.append((_date_str, _dt, _cat, _note, amt))

	for date_str, dt, category, note, amt, share, reason in candidate_shared:
		include = True
		if cutoff_date and dt is not None:
			include = dt.date() >= cutoff_date
		elif cutoff_date and dt is None:
			include = False
		if include:
			total_shared += share
			shared_rows.append((date_str, category, note, amt, share, reason))

	remaining = total_shared - total_paid_by_person

	data = {
		'totals': {
			'total_shared': str(total_shared),
			'total_paid_by_person': str(total_paid_by_person),
			'remaining': str(remaining),
			'person': person,
			'cutoff': str(cutoff_date) if cutoff_date else None,
		},
		'shared_rows': [
			{
				'date': d,
				'category': cat,
				'note': note,
				'amount': str(amt),
				'share': str(share),
				'reason': reason,
			}
			for d, cat, note, amt, share, reason in shared_rows
		],
		'applied_payments': [
			{'date': d, 'category': cat, 'note': note, 'amount': str(amt)}
			for d, dt, cat, note, amt in applied_payment_rows
		],
		'unapplied_payments': [
			{'date': d, 'category': cat, 'note': note, 'amount': str(amt)}
			for d, dt, cat, note, amt in unapplied_payment_rows
		],
	}

	return data

def main():
	base = os.path.dirname(__file__)
	files = list_transaction_files(base)

	parser = argparse.ArgumentParser(description='Tính tiền người cần trả từ file giao dịch')
	parser.add_argument('--file', '-f', help='Tên file CSV trong thư mục Transactions. Nếu không chỉ định sẽ hiển thị danh sách.')
	parser.add_argument('--person', '-p', default='Quân', help='Tên người (mặc định: Quân)')
	parser.add_argument('--paid-on', help='(Tùy chọn) Ngày Quân trả gần nhất theo định dạng YYYY-MM-DD. Nếu chỉ định, sẽ bắt đầu tính từ ngày hôm sau.')
	parser.add_argument('--start', help='(Tùy chọn) Ngày bắt đầu tính (YYYY-MM-DD). Nếu chỉ định, sẽ bắt đầu tính từ ngày này (bao gồm ngày).')
	parser.add_argument('--export', help='Các định dạng xuất, cách nhau bằng dấu phẩy. Hỗ trợ: csv,json,xlsx,pdf', default='')
	parser.add_argument('--outdir', help='Thư mục xuất file (mặc định: current dir)', default='.')
	args = parser.parse_args()

	chosen_file = args.file
	if not chosen_file:
		if not files:
			print('Không tìm thấy thư mục Transactions hoặc không có file nào bên trong.')
			return
		print('Các file trong thư mục Transactions:')
		for i, fn in enumerate(files, 1):
			print(f"  {i}. {fn}")
		try:
			sel = input('\nNhập số tương ứng để chọn file (hoặc Enter để hủy): ').strip()
		except EOFError:
			print('\nKhông nhận được đầu vào. Hủy.')
			return
		if sel == '':
			print('Hủy chọn file.')
			return
		try:
			idx = int(sel)
			if idx < 1 or idx > len(files):
				print('Số không hợp lệ.')
				return
			chosen_file = files[idx - 1]
			print(f'Chọn file: {chosen_file}')
		except ValueError:
			print('Giá trị nhập không phải số.')
			return

	csv_path = os.path.join(base, TRANSACTIONS_DIR, chosen_file)
	if not os.path.exists(csv_path):
		print('Không tìm thấy file CSV:', csv_path)
		return

	# Use computation function
	data = compute_summary(csv_path, person=args.person, paid_on=args.paid_on, start=args.start)

	# Determine cutoff/start date. Priority:
	# 1) --start (inclusive)
	# 2) --paid-on (interpreted as last payment date; start = paid_on + 1 day)
	# 3) interactive prompt (if TTY)
	# 4) auto-detect from payments_by_person (last payment + 1 day)
	# Derive display totals
	cutoff_date = data['totals']['cutoff']
	total_shared = Decimal(data['totals']['total_shared'])
	total_paid_by_person = Decimal(data['totals']['total_paid_by_person'])
	remaining = Decimal(data['totals']['remaining'])

	shared_rows = [
		(r['date'], r['category'], r['note'], Decimal(r['amount']), Decimal(r['share']), r['reason'])
		for r in data['shared_rows']
	]
	applied_payment_rows = [
		(r['date'], None, r['category'], r['note'], Decimal(r['amount']))
		for r in data['applied_payments']
	]
	unapplied_payment_rows = [
		(r['date'], None, r['category'], r['note'], Decimal(r['amount']))
		for r in data['unapplied_payments']
	]

	def fmt(v: Decimal) -> str:
		v = v.quantize(Decimal('1'))
		s = f"{int(v):,}"
		return s.replace(',', '.')

	print('\n=== KẾT QUẢ TÍNH TOÁN ===')
	print('Tổng phần liên quan (ưu tiên note nếu có):', fmt(total_shared), 'VND')
	print(f'Tổng {args.person} đã trả (ghi chú/loan):', fmt(total_paid_by_person), 'VND')
	print(f'Số tiền {args.person} còn nợ tôi:', fmt(remaining), 'VND')

	if shared_rows:
		print('\n--- Chi tiết phần liên quan ---')
		for d, cat, note, amt, share, reason in shared_rows:
			print(d, '|', cat, '|', note or '', '| amount:', int(amt), '| share:', int(share), '| reason:', reason)

	if applied_payment_rows or unapplied_payment_rows:
		print('\n--- Chi tiết các khoản đã trả bởi người ---')
		if applied_payment_rows:
			print('\nCác khoản áp dụng cho kỳ hiện tại:')
			for d, dt, cat, note, amt in applied_payment_rows:
				print(d, '|', cat, '|', note or '', '| amount:', int(amt))
		if unapplied_payment_rows:
			print('\nCác khoản không áp dụng (trước cutoff):')
			for d, dt, cat, note, amt in unapplied_payment_rows:
				print(d, '|', cat, '|', note or '', '| amount:', int(amt))

	# Exporting
	if not args.export:
		opts = [
			('csv', 'CSV', True),
			('json', 'JSON', True),
			('md', 'Markdown', True),
			('xlsx', 'XLSX', HAS_OPENPYXL),
			('pdf', 'PDF', HAS_REPORTLAB),
		]
		print('\nBạn có thể xuất kết quả sang các định dạng sau:')
		for i, (_, label, avail) in enumerate(opts, 1):
			print(f"  {i}. {label}" + ('' if avail else ' (không khả dụng; cần cài đặt)'))
		try:
			sel = input('\nNhập số tương ứng để chọn (ví dụ: 1 hoặc 1,3). Enter để bỏ qua xuất file: ').strip()
		except EOFError:
			sel = ''
		selected = []
		if sel:
			parts = [p.strip() for p in sel.split(',') if p.strip()]
			for p in parts:
				try:
					idx = int(p)
				except Exception:
					continue
				if idx < 1 or idx > len(opts):
					continue
				key, label, avail = opts[idx-1]
				if not avail:
					print(f"Bỏ qua {label}: chưa cài gói cần thiết.")
					continue
				selected.append(key)
			args.export = ','.join(selected)

	exports = [x.strip().lower() for x in (args.export or '').split(',') if x.strip()]
	if exports:
		outdir = args.outdir
		if not os.path.isdir(outdir):
			try:
				os.makedirs(outdir, exist_ok=True)
			except Exception as e:
				print('Không thể tạo thư mục xuất:', e)
				return

		base_name = os.path.splitext(os.path.basename(chosen_file))[0]
		# reuse data from computation

		for fmt in exports:
			if fmt == 'csv':
				out_csv = os.path.join(outdir, f"{base_name}.{args.person}.summary.csv")
				try:
					with open(out_csv, 'w', encoding='utf-8', newline='') as fo:
						w = csv.writer(fo)
						w.writerow(['section', 'date', 'category', 'note', 'amount', 'share', 'reason'])
						for r in data['shared_rows']:
							w.writerow(['shared', r['date'], r['category'], r['note'], r['amount'], r['share'], r['reason']])
						w.writerow([])
						w.writerow(['applied_payment', 'date', 'category', 'note', 'amount'])
						for r in data['applied_payments']:
							w.writerow(['applied', r['date'], r['category'], r['note'], r['amount']])
						w.writerow([])
						w.writerow(['unapplied_payment', 'date', 'category', 'note', 'amount'])
						for r in data['unapplied_payments']:
							w.writerow(['unapplied', r['date'], r['category'], r['note'], r['amount']])
						w.writerow([])
						w.writerow(['total_shared', data['totals']['total_shared']])
						w.writerow(['total_paid_by_person', data['totals']['total_paid_by_person']])
						w.writerow(['remaining', data['totals']['remaining']])
				except Exception as e:
					print('Lỗi khi ghi CSV:', e)
				else:
					print('Đã xuất CSV ->', out_csv)

			elif fmt == 'json':
				out_json = os.path.join(outdir, f"{base_name}.{args.person}.summary.json")
				try:
					with open(out_json, 'w', encoding='utf-8') as fo:
						json.dump(data, fo, ensure_ascii=False, indent=2)
				except Exception as e:
					print('Lỗi khi ghi JSON:', e)
				else:
					print('Đã xuất JSON ->', out_json)

			elif fmt == 'md':
				out_md = os.path.join(outdir, f"{base_name}.{args.person}.summary.md")
				try:
					with open(out_md, 'w', encoding='utf-8') as mf:
						mf.write(f"# Summary for {args.person} - {base_name}\n\n")
						mf.write('## Totals\n')
						mf.write(f"- Total shared: {data['totals']['total_shared']}\n")
						mf.write(f"- Total paid by {args.person}: {data['totals']['total_paid_by_person']}\n")
						mf.write(f"- Remaining: {data['totals']['remaining']}\n\n")
						mf.write('## Shared rows\n\n')
						mf.write('| date | category | note | amount | share | reason |\n')
						mf.write('|---|---|---|---:|---:|---|\n')
						for r in data['shared_rows']:
							note_safe = (r['note'] or '').replace('|', '\\|')
							mf.write(f"| {r['date']} | {r['category']} | {note_safe} | {r['amount']} | {r['share']} | {r['reason']} |\n")
						mf.write('\n## Applied payments\n\n')
						mf.write('| date | category | note | amount |\n')
						mf.write('|---|---|---|---:|\n')
						for r in data['applied_payments']:
							note_safe = (r['note'] or '').replace('|', '\\|')
							mf.write(f"| {r['date']} | {r['category']} | {note_safe} | {r['amount']} |\n")
						mf.write('\n## Unapplied payments\n\n')
						mf.write('| date | category | note | amount |\n')
						mf.write('|---|---|---|---:|\n')
						for r in data['unapplied_payments']:
							note_safe = (r['note'] or '').replace('|', '\\|')
							mf.write(f"| {r['date']} | {r['category']} | {note_safe} | {r['amount']} |\n")
				except Exception as e:
					print('Lỗi khi ghi Markdown:', e)
				else:
					print('Đã xuất Markdown ->', out_md)

			elif fmt == 'xlsx':
				if not HAS_OPENPYXL:
					print('openpyxl không được cài đặt. Bỏ qua XLSX export. Cài bằng: pip install openpyxl')
					continue
				out_xlsx = os.path.join(outdir, f"{base_name}.{args.person}.summary.xlsx")
				try:
					wb = Workbook()
					ws = wb.active
					ws.title = 'shared'
					ws.append(['date', 'category', 'note', 'amount', 'share', 'reason'])
					for r in data['shared_rows']:
						ws.append([r['date'], r['category'], r['note'], r['amount'], r['share'], r['reason']])
					ws2 = wb.create_sheet('applied_payments')
					ws2.append(['date', 'category', 'note', 'amount'])
					for r in data['applied_payments']:
						ws2.append([r['date'], r['category'], r['note'], r['amount']])
					ws3 = wb.create_sheet('unapplied_payments')
					ws3.append(['date', 'category', 'note', 'amount'])
					for r in data['unapplied_payments']:
						ws3.append([r['date'], r['category'], r['note'], r['amount']])
					ws4 = wb.create_sheet('totals')
					ws4.append(['total_shared', data['totals']['total_shared']])
					ws4.append(['total_paid_by_person', data['totals']['total_paid_by_person']])
					ws4.append(['remaining', data['totals']['remaining']])
					wb.save(out_xlsx)
				except Exception as e:
					print('Lỗi khi ghi XLSX:', e)
				else:
					print('Đã xuất XLSX ->', out_xlsx)

			elif fmt == 'pdf':
				if not HAS_REPORTLAB:
					print('reportlab không được cài. Bỏ qua PDF export. Cài bằng: pip install reportlab')
					continue
				out_pdf = os.path.join(outdir, f"{base_name}.{args.person}.summary.pdf")
				try:
					c = canvas.Canvas(out_pdf, pagesize=A4)
					w, h = A4
					y = h - 40
					c.setFont('Helvetica-Bold', 12)
					c.drawString(40, y, f"Summary for {args.person} - {base_name}")
					y -= 24
					c.setFont('Helvetica', 10)
					c.drawString(40, y, f"Totals: shared={data['totals']['total_shared']} paid={data['totals']['total_paid_by_person']} remaining={data['totals']['remaining']}")
					y -= 24
					c.drawString(40, y, 'Shared rows:')
					y -= 18
					for r in data['shared_rows']:
						line = f"{r['date']} | {r['category']} | {r['note'] or ''} | amt:{r['amount']} | share:{r['share']}"
						c.drawString(40, y, line[:120])
						y -= 14
						if y < 80:
							c.showPage()
							y = h - 40
					c.save()
				except Exception as e:
					print('Lỗi khi ghi PDF:', e)
				else:
					print('Đã xuất PDF ->', out_pdf)


if __name__ == '__main__':
	main()

