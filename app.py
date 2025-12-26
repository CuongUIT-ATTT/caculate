import os
import io
import csv
import json
from decimal import Decimal
import importlib.util
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, send_file
from werkzeug.utils import secure_filename
import tempfile

BASE_DIR = os.path.dirname(__file__)
TRANSACTIONS_DIR = os.path.join(BASE_DIR, 'Transactions')
UPLOADS_DIR = os.path.join(BASE_DIR, 'Uploads')

# Dynamically load compute_summary from Caculate-auto.py (hyphenated filename)
_script_path = os.path.join(BASE_DIR, 'Caculate-auto.py')
spec = importlib.util.spec_from_file_location("calc_module", _script_path)
calc_module = importlib.util.module_from_spec(spec)
spec.loader.exec_module(calc_module)
compute_summary = getattr(calc_module, 'compute_summary')

# Optional export libs
try:
    from openpyxl import Workbook
    HAS_OPENPYXL = True
except Exception:
    HAS_OPENPYXL = False

try:
    from reportlab.lib.pagesizes import A4
    from reportlab.pdfgen import canvas
    HAS_REPORTLAB = True
except Exception:
    HAS_REPORTLAB = False

app = Flask(__name__)


def list_transaction_files():
    if not os.path.isdir(TRANSACTIONS_DIR):
        return []
    files = [f for f in os.listdir(TRANSACTIONS_DIR) if os.path.isfile(os.path.join(TRANSACTIONS_DIR, f))]
    files.sort()
    return files


@app.route('/', methods=['GET'])
def index():
    files = list_transaction_files()
    return render_template('index.html', files=files)


@app.route('/compute', methods=['POST'])
def compute():
    # Ensure uploads folder exists
    os.makedirs(UPLOADS_DIR, exist_ok=True)

    file_name = (request.form.get('file') or '').strip()
    person = (request.form.get('person') or 'Quân').strip()
    paid_on = (request.form.get('paid_on') or None)
    start = (request.form.get('start') or None)

    uploaded_flag = False
    upload_file = request.files.get('upload')
    csv_path = None

    if upload_file and upload_file.filename:
        # Save uploaded CSV
        uploaded_flag = True
        fn = secure_filename(upload_file.filename)
        ts = datetime.now().strftime('%Y%m%d-%H%M%S')
        saved_name = f"{ts}_{fn}"
        csv_path = os.path.join(UPLOADS_DIR, saved_name)
        upload_file.save(csv_path)
        file_name = saved_name
    else:
        if not file_name:
            return redirect(url_for('index'))
        csv_path = os.path.join(TRANSACTIONS_DIR, file_name)
        if not os.path.exists(csv_path):
            return redirect(url_for('index'))

    params = {
        'file': file_name,
        'uploaded': '1' if uploaded_flag else '0',
        'person': person,
        'paid_on': paid_on or '',
        'start': start or '',
    }
    # Redirect to shareable GET page
    return redirect(url_for('result_view', **params))


@app.route('/result', methods=['GET'])
def result_view():
    file_name = (request.args.get('file') or '').strip()
    uploaded_flag = request.args.get('uploaded') or '0'
    person = (request.args.get('person') or 'Quân').strip()
    paid_on = (request.args.get('paid_on') or None)
    start = (request.args.get('start') or None)

    if not file_name:
        return redirect(url_for('index'))

    csv_path = _resolve_path(file_name, uploaded_flag)
    if not os.path.exists(csv_path):
        return redirect(url_for('index'))

    data = compute_summary(csv_path, person=person, paid_on=paid_on, start=start)

    # Convert to display-friendly values
    def fmt_int_str(s):
        v = Decimal(s).quantize(Decimal('1'))
        return f"{int(v):,}".replace(',', '.')

    totals = {
        'total_shared': fmt_int_str(data['totals']['total_shared']),
        'total_paid_by_person': fmt_int_str(data['totals']['total_paid_by_person']),
        'remaining': fmt_int_str(data['totals']['remaining']),
        'person': data['totals']['person'],
        'cutoff': data['totals']['cutoff'],
        'base_name': os.path.splitext(os.path.basename(file_name))[0],
    }

    available_exports = ['csv', 'json', 'md']
    if HAS_OPENPYXL:
        available_exports.append('xlsx')
    if HAS_REPORTLAB:
        available_exports.append('pdf')

    # Parameters (also keep filters for sharing)
    params = {
        'file': file_name,
        'uploaded': uploaded_flag,
        'person': person,
        'paid_on': paid_on or '',
        'start': start or '',
        'q': (request.args.get('q') or ''),
        'start_date': (request.args.get('start_date') or ''),
        'end_date': (request.args.get('end_date') or ''),
        'category': (request.args.get('category') or ''),
    }

    # Collect categories for filter
    categories = set()
    for r in data['shared_rows']:
        if r.get('category'):
            categories.add(r['category'])
    for r in data['applied_payments']:
        if r.get('category'):
            categories.add(r['category'])
    for r in data['unapplied_payments']:
        if r.get('category'):
            categories.add(r['category'])
    categories = sorted(categories)

    return render_template(
        'result.html',
        totals=totals,
        shared_rows=data['shared_rows'],
        applied_payments=data['applied_payments'],
        unapplied_payments=data['unapplied_payments'],
        available_exports=available_exports,
        params=params,
        categories=categories,
    )


def _resolve_path(file_name: str, uploaded_flag: str) -> str:
    if uploaded_flag == '1':
        return os.path.join(UPLOADS_DIR, file_name)
    return os.path.join(TRANSACTIONS_DIR, file_name)


@app.route('/export/<fmt>', methods=['GET'])
def export(fmt: str):
    fmt = (fmt or '').lower()
    file_name = (request.args.get('file') or '').strip()
    uploaded_flag = request.args.get('uploaded') or '0'
    person = (request.args.get('person') or 'Quân').strip()
    paid_on = (request.args.get('paid_on') or None)
    start = (request.args.get('start') or None)
    q = (request.args.get('q') or '').strip().lower()
    start_date = (request.args.get('start_date') or '').strip()
    end_date = (request.args.get('end_date') or '').strip()
    category = (request.args.get('category') or '').strip()

    if not file_name:
        return redirect(url_for('index'))

    csv_path = _resolve_path(file_name, uploaded_flag)
    if not os.path.exists(csv_path):
        return redirect(url_for('index'))

    data = compute_summary(csv_path, person=person, paid_on=paid_on, start=start)
    base_name = os.path.splitext(os.path.basename(file_name))[0]

    # Filtering helpers
    def in_range(date_str):
        if not (start_date or end_date):
            return True
        try:
            dt = datetime.fromisoformat(date_str.split('+')[0])
            d = dt.date()
        except Exception:
            return False
        if start_date:
            try:
                s = datetime.strptime(start_date, '%Y-%m-%d').date()
                if d < s:
                    return False
            except Exception:
                pass
        if end_date:
            try:
                e = datetime.strptime(end_date, '%Y-%m-%d').date()
                if d > e:
                    return False
            except Exception:
                pass
        return True

    def matches_q_row(row_dict):
        if not q:
            return True
        for key in ('date','category','note','amount','share','reason'):
            val = str(row_dict.get(key, '')).lower()
            if q in val:
                return True
        return False

    def matches_q_pay(row_dict):
        if not q:
            return True
        for key in ('date','category','note','amount'):
            val = str(row_dict.get(key, '')).lower()
            if q in val:
                return True
        return False

    def matches_category(cat):
        if not category:
            return True
        return (cat or '').lower() == category.lower()

    # Apply filters
    filtered = {
        'shared_rows': [r for r in data['shared_rows'] if in_range(r['date']) and matches_category(r['category']) and matches_q_row(r)],
        'applied_payments': [r for r in data['applied_payments'] if in_range(r['date']) and matches_category(r['category']) and matches_q_pay(r)],
        'unapplied_payments': [r for r in data['unapplied_payments'] if in_range(r['date']) and matches_category(r['category']) and matches_q_pay(r)],
        'totals': data['totals'],
    }

    if fmt == 'csv':
        bio = io.StringIO()
        w = csv.writer(bio)
        w.writerow(['section', 'date', 'category', 'note', 'amount', 'share', 'reason'])
        for r in filtered['shared_rows']:
            w.writerow(['shared', r['date'], r['category'], r['note'], r['amount'], r['share'], r['reason']])
        w.writerow([])
        w.writerow(['applied_payment', 'date', 'category', 'note', 'amount'])
        for r in filtered['applied_payments']:
            w.writerow(['applied', r['date'], r['category'], r['note'], r['amount']])
        w.writerow([])
        w.writerow(['unapplied_payment', 'date', 'category', 'note', 'amount'])
        for r in filtered['unapplied_payments']:
            w.writerow(['unapplied', r['date'], r['category'], r['note'], r['amount']])
        w.writerow([])
        w.writerow(['total_shared', filtered['totals']['total_shared']])
        w.writerow(['total_paid_by_person', filtered['totals']['total_paid_by_person']])
        w.writerow(['remaining', filtered['totals']['remaining']])
        bytes_io = io.BytesIO(bio.getvalue().encode('utf-8'))
        return send_file(bytes_io, as_attachment=True, download_name=f"{base_name}.{person}.summary.csv", mimetype='text/csv')

    if fmt == 'json':
        content = json.dumps(filtered, ensure_ascii=False, indent=2)
        bytes_io = io.BytesIO(content.encode('utf-8'))
        return send_file(bytes_io, as_attachment=True, download_name=f"{base_name}.{person}.summary.json", mimetype='application/json')

    if fmt == 'md':
        lines = []
        lines.append(f"# Summary for {person} - {base_name}\n\n")
        lines.append('## Totals\n')
        lines.append(f"- Total shared: {filtered['totals']['total_shared']}\n")
        lines.append(f"- Total paid by {person}: {filtered['totals']['total_paid_by_person']}\n")
        lines.append(f"- Remaining: {filtered['totals']['remaining']}\n\n")
        lines.append('## Shared rows\n\n')
        lines.append('| date | category | note | amount | share | reason |\n')
        lines.append('|---|---|---|---:|---:|---|\n')
        for r in filtered['shared_rows']:
            note = (r['note'] or '').replace('|', '\\|')
            lines.append(f"| {r['date']} | {r['category']} | {note} | {r['amount']} | {r['share']} | {r['reason']} |\n")
        lines.append('\n## Applied payments\n\n')
        lines.append('| date | category | note | amount |\n')
        lines.append('|---|---|---|---:|\n')
        for r in filtered['applied_payments']:
            note = (r['note'] or '').replace('|', '\\|')
            lines.append(f"| {r['date']} | {r['category']} | {note} | {r['amount']} |\n")
        lines.append('\n## Unapplied payments\n\n')
        lines.append('| date | category | note | amount |\n')
        lines.append('|---|---|---|---:|\n')
        for r in filtered['unapplied_payments']:
            note = (r['note'] or '').replace('|', '\\|')
            lines.append(f"| {r['date']} | {r['category']} | {note} | {r['amount']} |\n")
        bytes_io = io.BytesIO(''.join(lines).encode('utf-8'))
        return send_file(bytes_io, as_attachment=True, download_name=f"{base_name}.{person}.summary.md", mimetype='text/markdown')

    if fmt == 'xlsx' and HAS_OPENPYXL:
        wb = Workbook()
        ws = wb.active
        ws.title = 'shared'
        ws.append(['date', 'category', 'note', 'amount', 'share', 'reason'])
        for r in filtered['shared_rows']:
            ws.append([r['date'], r['category'], r['note'], r['amount'], r['share'], r['reason']])
        ws2 = wb.create_sheet('applied_payments')
        ws2.append(['date', 'category', 'note', 'amount'])
        for r in filtered['applied_payments']:
            ws2.append([r['date'], r['category'], r['note'], r['amount']])
        ws3 = wb.create_sheet('unapplied_payments')
        ws3.append(['date', 'category', 'note', 'amount'])
        for r in filtered['unapplied_payments']:
            ws3.append([r['date'], r['category'], r['note'], r['amount']])
        ws4 = wb.create_sheet('totals')
        ws4.append(['total_shared', filtered['totals']['total_shared']])
        ws4.append(['total_paid_by_person', filtered['totals']['total_paid_by_person']])
        ws4.append(['remaining', filtered['totals']['remaining']])
        bytes_io = io.BytesIO()
        wb.save(bytes_io)
        bytes_io.seek(0)
        return send_file(bytes_io, as_attachment=True, download_name=f"{base_name}.{person}.summary.xlsx", mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    if fmt == 'pdf' and HAS_REPORTLAB:
        bytes_io = io.BytesIO()
        c = canvas.Canvas(bytes_io, pagesize=A4)
        w, h = A4
        y = h - 40
        c.setFont('Helvetica-Bold', 12)
        c.drawString(40, y, f"Summary for {person} - {base_name}")
        y -= 24
        c.setFont('Helvetica', 10)
        c.drawString(40, y, f"Totals: shared={filtered['totals']['total_shared']} paid={filtered['totals']['total_paid_by_person']} remaining={filtered['totals']['remaining']}")
        y -= 24
        c.drawString(40, y, 'Shared rows:')
        y -= 18
        for r in filtered['shared_rows']:
            line = f"{r['date']} | {r['category']} | {r['note'] or ''} | amt:{r['amount']} | share:{r['share']}"
            c.drawString(40, y, line[:120])
            y -= 14
            if y < 80:
                c.showPage()
                y = h - 40
        c.save()
        bytes_io.seek(0)
        return send_file(bytes_io, as_attachment=True, download_name=f"{base_name}.{person}.summary.pdf", mimetype='application/pdf')

    return redirect(url_for('index'))


@app.route('/pdf-to-csv', methods=['POST'])
def pdf_to_csv():
    # Ensure uploads folder exists
    os.makedirs(UPLOADS_DIR, exist_ok=True)

    pdf_file = request.files.get('pdf')
    person = (request.form.get('person') or 'Quân').strip()
    paid_on = (request.form.get('paid_on') or None)
    start = (request.form.get('start') or None)
    if not pdf_file or not pdf_file.filename.lower().endswith('.pdf'):
        return redirect(url_for('index'))

    fn = secure_filename(pdf_file.filename)
    ts = datetime.now().strftime('%Y%m%d-%H%M%S')
    saved_pdf_name = f"{ts}_{fn}"
    saved_pdf_path = os.path.join(UPLOADS_DIR, saved_pdf_name)
    pdf_file.save(saved_pdf_path)

    base_name = os.path.splitext(os.path.basename(fn))[0]
    out_csv_name = f"{ts}_{base_name}.extracted.csv"
    out_csv_path = os.path.join(UPLOADS_DIR, out_csv_name)

    # Lazy import to avoid hard dependency unless used
    try:
        from pdf_to_csv import extract_tables_to_structured_csv
    except Exception:
        extract_tables_to_structured_csv = None

    if extract_tables_to_structured_csv is None:
        # As a minimal fallback, use the raw extractor and then wrap to the schema with basic mapping
        try:
            from pdf_to_csv import extract_tables_to_csv
            extract_tables_to_csv(saved_pdf_path, out_csv_path)
        except Exception:
            return redirect(url_for('index'))
    else:
        try:
            extract_tables_to_structured_csv(saved_pdf_path, out_csv_path)
        except Exception:
            return redirect(url_for('index'))

    # Redirect straight to results to allow immediate computation/view
    params = {
        'file': out_csv_name,
        'uploaded': '1',
        'person': person,
        'paid_on': paid_on or '',
        'start': start or '',
    }
    return redirect(url_for('result_view', **params))


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
